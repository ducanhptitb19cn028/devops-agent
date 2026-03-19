"""
Export ML training results, model metrics, and datasets to Excel.

Usage:
    python -m ml-models.export_to_excel
    python ml-models/export_to_excel.py [--output results.xlsx] [--include-data]
"""
import argparse
import json
import os
from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

BASE_DIR = Path(__file__).parent
ARTIFACTS_DIR = BASE_DIR / "artifacts"
DATA_DIR = BASE_DIR / "data"
TRAINED_MODELS_DIR = BASE_DIR / "trained_models"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _style_header_row(ws, row: int, fill_hex: str = "1F4E79"):
    fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")
    font = Font(bold=True, color="FFFFFF", size=11)
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)


def _thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def _write_df(ws, df: pd.DataFrame, start_row: int = 1):
    """Write a DataFrame to a worksheet starting at start_row."""
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=col_name)
        cell.border = _thin_border()

    _style_header_row(ws, start_row)

    for row_idx, row in enumerate(df.itertuples(index=False), start_row + 1):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = _thin_border()
            cell.alignment = Alignment(horizontal="left")
            if isinstance(value, float):
                cell.number_format = "0.0000"


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _sheet_summary(wb, all_results: dict):
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False

    title_cell = ws["A1"]
    title_cell.value = "ML Training Results — TraceFlix DevOps Agent"
    title_cell.font = Font(bold=True, size=14, color="1F4E79")
    ws.merge_cells("A1:G1")

    ts_cell = ws["A2"]
    ts_cell.value = f"Exported: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ts_cell.font = Font(italic=True, size=10, color="595959")
    ws.merge_cells("A2:G2")

    rows = [
        ["Model", "Training Time (s)", "Key Metric 1", "Value 1", "Key Metric 2", "Value 2", "Status"],
    ]

    for name, res in all_results.items():
        if name == "anomaly":
            rows.append([
                "Anomaly Detection",
                round(res.get("training_time_s", 0), 1),
                "Test F1", round(res["test"]["test_f1"], 4),
                "Test AUC", round(res["test"]["test_auc"], 4),
                "✔ Trained",
            ])
        elif name == "forecasting":
            rows.append([
                "Forecasting (LSTM+Attention)",
                round(res.get("training_time_s", 0), 1),
                "Val MAE", round(res.get("val_mae", 0), 2),
                "Val RMSE", round(res.get("val_rmse", 0), 2),
                "✔ Trained",
            ])
        elif name == "root_cause":
            rows.append([
                "Root Cause (XGBoost)",
                round(res.get("training_time_s", 0), 1),
                "Test Accuracy", round(res.get("test_accuracy", 0), 4),
                "Test F1 (weighted)", round(res.get("test_f1_weighted", 0), 4),
                "✔ Trained",
            ])
        elif name == "log_clustering":
            rows.append([
                "Log Clustering (HDBSCAN)",
                round(res.get("training_time_s", 0), 1),
                "Silhouette Score", round(res.get("silhouette_score", 0), 4),
                "N Clusters", res.get("n_clusters", 0),
                "✔ Trained",
            ])

    for r_idx, row in enumerate(rows, 4):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = _thin_border()
            cell.alignment = Alignment(horizontal="center")
            if isinstance(val, float):
                cell.number_format = "0.0000"

    _style_header_row(ws, 4)
    _auto_width(ws)
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[4].height = 20


def _sheet_anomaly(wb, res: dict):
    ws = wb.create_sheet("Anomaly Detection")
    ws.sheet_view.showGridLines = False

    # --- Isolation Forest metrics ---
    ws["A1"].value = "Isolation Forest"
    ws["A1"].font = Font(bold=True, size=12, color="1F4E79")

    df_if = pd.DataFrame([{
        "Metric": "Precision",
        "Value": res["isolation_forest"]["precision"],
    }, {
        "Metric": "Recall",
        "Value": res["isolation_forest"]["recall"],
    }, {
        "Metric": "F1",
        "Value": res["isolation_forest"]["f1"],
    }])
    _write_df(ws, df_if, start_row=2)

    # --- LSTM Autoencoder metrics ---
    ws["A7"].value = "LSTM Autoencoder"
    ws["A7"].font = Font(bold=True, size=12, color="1F4E79")

    df_lstm = pd.DataFrame([{
        "Metric": k.replace("_", " ").title(),
        "Value": v,
    } for k, v in res["lstm_autoencoder"].items() if k != "threshold"])
    df_lstm = pd.concat([df_lstm, pd.DataFrame([{
        "Metric": "Threshold",
        "Value": float(res["lstm_autoencoder"]["threshold"]),
    }])], ignore_index=True)
    _write_df(ws, df_lstm, start_row=8)

    # --- Ensemble Test metrics ---
    ws["A15"].value = "Ensemble Test Results"
    ws["A15"].font = Font(bold=True, size=12, color="1F4E79")

    df_test = pd.DataFrame([{
        "Metric": k.replace("test_", "").replace("_", " ").title(),
        "Value": v,
    } for k, v in res["test"].items()])
    _write_df(ws, df_test, start_row=16)

    # --- Bar chart for test metrics ---
    chart = BarChart()
    chart.type = "col"
    chart.title = "Anomaly Detection — Ensemble Test Metrics"
    chart.y_axis.title = "Score"
    chart.x_axis.title = "Metric"
    chart.shape = 4
    data_ref = Reference(ws, min_col=2, min_row=16, max_row=16 + len(df_test))
    cats_ref = Reference(ws, min_col=1, min_row=17, max_row=16 + len(df_test))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.width = 18
    chart.height = 12
    ws.add_chart(chart, "D2")

    _auto_width(ws)


def _sheet_forecasting(wb, res: dict):
    ws = wb.create_sheet("Forecasting")
    ws.sheet_view.showGridLines = False

    # Overall metrics
    ws["A1"].value = "Overall Metrics"
    ws["A1"].font = Font(bold=True, size=12, color="1F4E79")

    df_overall = pd.DataFrame([
        {"Split": "Train", "MAE": res["train_mae"], "RMSE": res["train_rmse"]},
        {"Split": "Val",   "MAE": res["val_mae"],   "RMSE": res["val_rmse"]},
    ])
    _write_df(ws, df_overall, start_row=2)

    # Per-feature MAE
    ws["A6"].value = "Per-Feature MAE"
    ws["A6"].font = Font(bold=True, size=12, color="1F4E79")

    features = list(res["train_per_feature_mae"].keys())
    df_feat = pd.DataFrame({
        "Feature": features,
        "Train MAE": [res["train_per_feature_mae"][f] for f in features],
        "Val MAE":   [res["val_per_feature_mae"][f] for f in features],
    })
    _write_df(ws, df_feat, start_row=7)

    # Bar chart
    chart = BarChart()
    chart.type = "col"
    chart.title = "Per-Feature MAE (Train vs Val)"
    chart.y_axis.title = "MAE"
    chart.x_axis.title = "Feature"
    data_ref = Reference(ws, min_col=2, min_row=7, max_col=3, max_row=7 + len(df_feat))
    cats_ref = Reference(ws, min_col=1, min_row=8, max_row=7 + len(df_feat))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.width = 22
    chart.height = 14
    ws.add_chart(chart, "E2")

    _auto_width(ws)


def _sheet_root_cause(wb, res: dict):
    ws = wb.create_sheet("Root Cause")
    ws.sheet_view.showGridLines = False

    ws["A1"].value = "Root Cause Classification (XGBoost)"
    ws["A1"].font = Font(bold=True, size=12, color="1F4E79")

    df = pd.DataFrame([
        {"Metric": "Train Accuracy",       "Value": res["train_accuracy"]},
        {"Metric": "Train F1 (weighted)",  "Value": res["train_f1_weighted"]},
        {"Metric": "CV F1 Mean (5-fold)",  "Value": res["cv_f1_mean"]},
        {"Metric": "CV F1 Std",            "Value": res["cv_f1_std"]},
        {"Metric": "Test Accuracy",        "Value": res["test_accuracy"]},
        {"Metric": "Test F1 (weighted)",   "Value": res["test_f1_weighted"]},
        {"Metric": "N Classes",            "Value": res["n_classes"]},
        {"Metric": "N Samples (original)", "Value": res["n_samples"]},
        {"Metric": "N Samples (SMOTE)",    "Value": res["n_samples_resampled"]},
        {"Metric": "Training Time (s)",    "Value": res["training_time_s"]},
    ])
    _write_df(ws, df, start_row=2)

    # Feature importance (if available)
    feat_imp_path = TRAINED_MODELS_DIR / "root_cause" / "feature_importance.json"
    if feat_imp_path.exists():
        with open(feat_imp_path) as f:
            feat_imp = json.load(f)

        ws["A14"].value = "Feature Importances (Top 20)"
        ws["A14"].font = Font(bold=True, size=12, color="1F4E79")

        sorted_feats = sorted(feat_imp.items(), key=lambda x: x[1], reverse=True)[:20]
        df_fi = pd.DataFrame(sorted_feats, columns=["Feature", "Importance"])
        _write_df(ws, df_fi, start_row=15)

        chart = BarChart()
        chart.type = "bar"
        chart.title = "Top 20 Feature Importances"
        chart.x_axis.title = "Importance"
        data_ref = Reference(ws, min_col=2, min_row=15, max_row=15 + len(df_fi))
        cats_ref = Reference(ws, min_col=1, min_row=16, max_row=15 + len(df_fi))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.width = 22
        chart.height = 16
        ws.add_chart(chart, "D2")

    _auto_width(ws)


def _sheet_log_clustering(wb, res: dict):
    ws = wb.create_sheet("Log Clustering")
    ws.sheet_view.showGridLines = False

    ws["A1"].value = "Log Clustering (SBERT + UMAP + HDBSCAN)"
    ws["A1"].font = Font(bold=True, size=12, color="1F4E79")

    df = pd.DataFrame([
        {"Metric": "N Clusters",              "Value": res["n_clusters"]},
        {"Metric": "N Noise Points",          "Value": res["n_noise_points"]},
        {"Metric": "Silhouette Score",        "Value": res["silhouette_score"]},
        {"Metric": "Adjusted Rand Index",     "Value": res["adjusted_rand_index"]},
        {"Metric": "Normalised Mutual Info",  "Value": res["normalised_mutual_info"]},
        {"Metric": "Homogeneity",             "Value": res["homogeneity"]},
        {"Metric": "Completeness",            "Value": res["completeness"]},
        {"Metric": "V-Measure",               "Value": res["v_measure"]},
        {"Metric": "Training Time (s)",       "Value": res["training_time_s"]},
    ])
    _write_df(ws, df, start_row=2)

    # Cluster patterns (if available)
    patterns_path = TRAINED_MODELS_DIR / "log_clustering" / "cluster_patterns.json"
    if patterns_path.exists():
        with open(patterns_path) as f:
            patterns = json.load(f)

        ws["A13"].value = "Cluster Patterns (Sample — up to 50)"
        ws["A13"].font = Font(bold=True, size=12, color="1F4E79")

        rows = []
        for cluster_id, info in list(patterns.items())[:50]:
            rows.append({
                "Cluster ID": cluster_id,
                "Size": info.get("size", ""),
                "Representative Message": str(info.get("representative", ""))[:200],
                "Severity": info.get("severity", ""),
            })
        if rows:
            df_pat = pd.DataFrame(rows)
            _write_df(ws, df_pat, start_row=14)

    _auto_width(ws)


def _sheet_dataset_info(wb):
    ws = wb.create_sheet("Dataset Info")
    ws.sheet_view.showGridLines = False

    ws["A1"].value = "Dataset Information"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")

    info = [
        ["Source", "Type", "File", "Description", "Records (approx)"],
        ["Synthetic Generator", "Metrics",
         "ml-models/data/metrics_raw.parquet",
         "Time-series metrics for 3 microservices (movie, actor, review). "
         "8 features: request_rate, error_rate, latency_p50/p99, jvm_heap_used, jvm_gc_pause, cpu_usage, memory_usage.",
         "~270,000 rows (4500 windows × 60 timesteps)"],
        ["Synthetic Generator", "Windowed Features",
         "ml-models/data/windows_features.parquet",
         "Statistical features extracted from metric windows: mean, std, min, max, trend, rate-of-change per feature.",
         "~4,500 rows"],
        ["Synthetic Generator", "Logs",
         "ml-models/data/logs_training.parquet",
         "20,000 synthetic log entries from Spring Boot services. "
         "24 log templates, severities: INFO/WARN/ERROR/FATAL. Includes cluster labels.",
         "20,000 rows"],
        ["Real Data (optional)", "All types",
         "ml-models/data/real/<timestamp>/",
         "Fetched from backend API: /api/metrics, /api/logs, /api/traces. "
         "Collected via collect_real_data.py with configurable lookback window.",
         "Variable"],
    ]

    for r_idx, row in enumerate(info, 3):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = _thin_border()
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    _style_header_row(ws, 3)

    # Anomaly types
    ws["A10"].value = "Injected Anomaly Patterns"
    ws["A10"].font = Font(bold=True, size=12, color="1F4E79")

    anomalies = [
        ["Anomaly Type", "Description"],
        ["memory_leak", "Gradually increasing jvm_heap_used"],
        ["cpu_saturation", "cpu_usage spikes to near-max"],
        ["downstream_timeout", "latency_p99 and error_rate increase together"],
        ["config_error", "High error_rate, normal latency"],
        ["jvm_gc_pressure", "Elevated jvm_gc_pause_seconds"],
        ["request_spike", "Sharp request_rate increase"],
        ["deployment_regression", "Gradual multi-metric degradation post-deploy"],
    ]
    for r_idx, row in enumerate(anomalies, 11):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = _thin_border()
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    _style_header_row(ws, 11)
    _auto_width(ws)
    ws.column_dimensions["D"].width = 60
    ws.column_dimensions["E"].width = 30


def _sheet_parquet(wb, name: str, path: Path, max_rows: int = 1000):
    """Load a parquet file and write it as a sheet."""
    if not path.exists():
        return
    try:
        df = pd.read_parquet(path)
    except Exception as e:
        ws = wb.create_sheet(name)
        ws["A1"].value = f"Could not load {path.name}: {e}"
        return

    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False

    ws["A1"].value = f"{path.name}  ({len(df):,} total rows — showing first {min(len(df), max_rows):,})"
    ws["A1"].font = Font(bold=True, size=11, color="1F4E79")

    sample = df.head(max_rows).reset_index(drop=True)
    # Convert any non-serialisable columns
    for col in sample.select_dtypes(include=["datetime64[ns]", "datetime64[ns, UTC]"]).columns:
        sample[col] = sample[col].astype(str)

    _write_df(ws, sample, start_row=3)
    _auto_width(ws)


# ---------------------------------------------------------------------------
# Main entry-point
# ---------------------------------------------------------------------------

def export(output_path: str = None, include_data: bool = False):
    # Find the latest training results
    result_files = sorted(ARTIFACTS_DIR.glob("training_results_*.json"), reverse=True)
    full_results_path = ARTIFACTS_DIR / "full_results.json"

    if not result_files and not full_results_path.exists():
        raise FileNotFoundError(
            f"No training result JSON files found in {ARTIFACTS_DIR}. "
            "Run `python -m pipeline.train_all` first."
        )

    results_path = result_files[0] if result_files else full_results_path
    with open(results_path) as f:
        results = json.load(f)

    print(f"Loaded results from: {results_path}")

    if output_path is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = str(ARTIFACTS_DIR / f"ml_results_{ts}.xlsx")

    # Create workbook
    import openpyxl
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    # Build sheets
    _sheet_summary(wb, results)

    if "anomaly" in results:
        _sheet_anomaly(wb, results["anomaly"])
    if "forecasting" in results:
        _sheet_forecasting(wb, results["forecasting"])
    if "root_cause" in results:
        _sheet_root_cause(wb, results["root_cause"])
    if "log_clustering" in results:
        _sheet_log_clustering(wb, results["log_clustering"])

    _sheet_dataset_info(wb)

    if include_data:
        _sheet_parquet(wb, "Data — Windows Features",
                       DATA_DIR / "windows_features.parquet", max_rows=500)
        _sheet_parquet(wb, "Data — Metrics Raw",
                       DATA_DIR / "metrics_raw.parquet", max_rows=500)
        _sheet_parquet(wb, "Data — Logs",
                       DATA_DIR / "logs_training.parquet", max_rows=500)

    wb.save(output_path)
    print(f"Exported → {output_path}")
    return output_path


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Export ML results to Excel")
    parser.add_argument("--output", default=None,
                        help="Output .xlsx path (default: artifacts/ml_results_<ts>.xlsx)")
    parser.add_argument("--include-data", action="store_true",
                        help="Also export parquet data samples as sheets (slow for large files)")
    args = parser.parse_args()
    export(output_path=args.output, include_data=args.include_data)
