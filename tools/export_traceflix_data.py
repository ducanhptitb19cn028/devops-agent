"""
Export live TraceFlix data from the backend API to Excel.

Fetches: logs, metrics, traces, K8s events, AI analyses, TSDB trends, stats summary.

Usage:
    python tools/export_traceflix_data.py
    python tools/export_traceflix_data.py --url http://localhost:8000 --since 120 --output data.xlsx

Requirements:
    pip install requests openpyxl pandas
"""
import argparse
import json
from datetime import datetime
from pathlib import Path

import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

DEFAULT_URL = "http://localhost:8000"
OUT_DIR = Path(__file__).parent.parent / "ml-models" / "artifacts"


# ── Styling helpers ──────────────────────────────────────────────────────────

def _thin():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def _header_row(ws, row: int, hex_color: str = "1F4E79"):
    fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
    font = Font(bold=True, color="FFFFFF", size=10)
    for cell in ws[row]:
        if cell.value is None:
            continue
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin()


def _auto_width(ws, max_width: int = 50):
    for col in ws.columns:
        w = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 3, max_width)


def _write_df(ws, df: pd.DataFrame, start_row: int = 1, header_color: str = "1F4E79"):
    if df.empty:
        ws.cell(row=start_row, column=1, value="No data returned.")
        return

    for col_i, name in enumerate(df.columns, 1):
        ws.cell(row=start_row, column=col_i, value=name)

    _header_row(ws, start_row, header_color)

    for row_i, row in enumerate(df.itertuples(index=False), start_row + 1):
        for col_i, val in enumerate(row, 1):
            cell = ws.cell(row=row_i, column=col_i, value=val)
            cell.border = _thin()
            cell.alignment = Alignment(vertical="top", wrap_text=False)
            if isinstance(val, float):
                cell.number_format = "0.0000"


# ── API fetch helpers ─────────────────────────────────────────────────────────

def _get(base_url: str, path: str, params: dict = None, timeout: int = 30) -> dict:
    try:
        r = requests.get(f"{base_url}{path}", params=params, timeout=timeout)
        r.raise_for_status()
        return r.json()
    except requests.exceptions.ConnectionError:
        print(f"  [ERROR] Cannot connect to {base_url}. Is the backend running?")
        print(f"          Run: make port-forward")
        return {}
    except Exception as e:
        print(f"  [WARN] {path} failed: {e}")
        return {}


def _parse_json_field(val, expected_type=list):
    """Parse a field that may be a JSON string or already the expected type."""
    if isinstance(val, expected_type):
        return val
    if isinstance(val, str):
        try:
            parsed = json.loads(val)
            if isinstance(parsed, expected_type):
                return parsed
        except (json.JSONDecodeError, ValueError):
            pass
    return expected_type()


def _flatten_labels(labels) -> str:
    """Convert a labels dict/str to a compact string."""
    if isinstance(labels, dict):
        return ", ".join(f"{k}={v}" for k, v in labels.items())
    return str(labels) if labels else ""


def _json_str(val) -> str:
    """Compact JSON string for JSONB fields."""
    if val is None:
        return ""
    if isinstance(val, (dict, list)):
        return json.dumps(val, separators=(",", ":"))
    return str(val)


# ── Sheet builders ────────────────────────────────────────────────────────────

def _sheet_summary(wb: Workbook, stats: dict, since_minutes: int):
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False

    ws["A1"].value = "TraceFlix Observability Data Export"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws.merge_cells("A1:D1")

    ws["A2"].value = f"Exported: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}   |   Window: last {since_minutes} minutes"
    ws["A2"].font = Font(italic=True, size=10, color="595959")
    ws.merge_cells("A2:D2")

    rows = [
        ["Data Type", "Count", "Notes"],
        ["Logs (total)", stats.get("total_logs", "—"), f"Last {since_minutes} min"],
        ["Logs (errors)", stats.get("total_errors", "—"), "Severity = ERROR"],
        ["Traces (total)", stats.get("total_traces", "—"), f"Last {since_minutes} min"],
        ["Traces (slow)", stats.get("slow_traces", "—"), "duration_ms > 500"],
        ["Traces (error)", stats.get("error_traces", "—"), "has_error = true"],
        ["K8s Events", stats.get("k8s_events", "—"), f"Last {since_minutes} min"],
    ]

    for r_i, row in enumerate(rows, 4):
        for c_i, val in enumerate(row, 1):
            cell = ws.cell(row=r_i, column=c_i, value=val)
            cell.border = _thin()
            cell.alignment = Alignment(horizontal="center")
    _header_row(ws, 4)

    # Errors by service
    errors_by_svc = stats.get("errors_by_service", [])
    if errors_by_svc:
        ws["A12"].value = "Errors by Service"
        ws["A12"].font = Font(bold=True, size=11, color="1F4E79")
        df_err = pd.DataFrame(errors_by_svc).rename(columns={"cnt": "error_count"})
        _write_df(ws, df_err, start_row=13, header_color="C00000")

    # Latency by service
    latency_by_svc = stats.get("latency_by_service", [])
    if latency_by_svc:
        ws["D12"].value = "Latency by Service (ms)"
        ws["D12"].font = Font(bold=True, size=11, color="1F4E79")
        df_lat = pd.DataFrame(latency_by_svc)
        _write_df(ws, df_lat, start_row=13, header_color="375623")

    _auto_width(ws)


def _sheet_logs(wb: Workbook, logs: list):
    ws = wb.create_sheet("Logs")
    ws.sheet_view.showGridLines = False

    rows = []
    for entry in logs:
        rows.append({
            "id": entry.get("id"),
            "timestamp": entry.get("timestamp", ""),
            "service": entry.get("service", ""),
            "severity": entry.get("severity", ""),
            "message": str(entry.get("message", ""))[:500],
            "namespace": entry.get("namespace", ""),
            "labels": _flatten_labels(entry.get("labels")),
            "created_at": entry.get("created_at", ""),
        })

    df = pd.DataFrame(rows) if rows else pd.DataFrame(
        columns=["id", "timestamp", "service", "severity", "message", "namespace", "labels", "created_at"]
    )
    _write_df(ws, df, start_row=1)
    ws.row_dimensions[1].height = 18

    # Color ERROR/WARN rows
    red_fill  = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
    warn_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    sev_col = list(df.columns).index("severity") + 1
    for row_i in range(2, len(df) + 2):
        sev = ws.cell(row=row_i, column=sev_col).value
        if sev == "ERROR":
            for cell in ws[row_i]:
                cell.fill = red_fill
        elif sev == "WARN":
            for cell in ws[row_i]:
                cell.fill = warn_fill

    _auto_width(ws)
    ws.column_dimensions[get_column_letter(list(df.columns).index("message") + 1)].width = 70


def _sheet_metrics(wb: Workbook, metrics: list):
    ws = wb.create_sheet("Metrics")
    ws.sheet_view.showGridLines = False

    rows = []
    for m in metrics:
        labels = m.get("labels", {})
        svc = labels.get("service_name", labels.get("job", "")) if isinstance(labels, dict) else ""
        rows.append({
            "id": m.get("id"),
            "timestamp": m.get("timestamp", ""),
            "metric_name": m.get("metric_name", ""),
            "service": svc,
            "value": m.get("value", ""),
            "namespace": m.get("namespace", ""),
            "labels": _flatten_labels(labels),
            "created_at": m.get("created_at", ""),
        })

    df = pd.DataFrame(rows) if rows else pd.DataFrame(
        columns=["id", "timestamp", "metric_name", "service", "value", "namespace", "labels", "created_at"]
    )
    _write_df(ws, df, start_row=1, header_color="375623")
    _auto_width(ws)


def _sheet_traces(wb: Workbook, traces: list):
    ws = wb.create_sheet("Traces")
    ws.sheet_view.showGridLines = False

    rows = []
    for t in traces:
        rows.append({
            "id": t.get("id"),
            "timestamp": t.get("timestamp", ""),
            "trace_id": t.get("trace_id", ""),
            "service": t.get("service", ""),
            "operation": t.get("operation", ""),
            "duration_ms": t.get("duration_ms"),
            "is_slow": t.get("is_slow"),
            "has_error": t.get("has_error"),
            "created_at": t.get("created_at", ""),
        })

    df = pd.DataFrame(rows) if rows else pd.DataFrame(
        columns=["id", "timestamp", "trace_id", "service", "operation",
                 "duration_ms", "is_slow", "has_error", "created_at"]
    )
    _write_df(ws, df, start_row=1, header_color="7030A0")

    # Highlight slow/error traces
    slow_fill  = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    error_fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
    cols = list(df.columns)
    slow_col  = cols.index("is_slow") + 1
    error_col = cols.index("has_error") + 1
    for row_i in range(2, len(df) + 2):
        is_slow  = ws.cell(row=row_i, column=slow_col).value
        has_err  = ws.cell(row=row_i, column=error_col).value
        if has_err:
            for cell in ws[row_i]: cell.fill = error_fill
        elif is_slow:
            for cell in ws[row_i]: cell.fill = slow_fill

    _auto_width(ws)


def _sheet_events(wb: Workbook, events: list):
    ws = wb.create_sheet("K8s Events")
    ws.sheet_view.showGridLines = False

    rows = []
    for e in events:
        rows.append({
            "id": e.get("id"),
            "timestamp": e.get("timestamp", ""),
            "source": e.get("source", ""),
            "pod": e.get("pod", ""),
            "reason": e.get("reason", ""),
            "event_type": e.get("event_type", ""),
            "message": str(e.get("message", ""))[:300],
            "namespace": e.get("namespace", ""),
            "created_at": e.get("created_at", ""),
        })

    df = pd.DataFrame(rows) if rows else pd.DataFrame(
        columns=["id", "timestamp", "source", "pod", "reason",
                 "event_type", "message", "namespace", "created_at"]
    )
    _write_df(ws, df, start_row=1, header_color="C55A11")

    warn_fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
    type_col = list(df.columns).index("event_type") + 1
    for row_i in range(2, len(df) + 2):
        if ws.cell(row=row_i, column=type_col).value == "Warning":
            for cell in ws[row_i]: cell.fill = warn_fill

    _auto_width(ws)
    ws.column_dimensions[get_column_letter(list(df.columns).index("message") + 1)].width = 60


def _sheet_analysis(wb: Workbook, analyses: list):
    ws = wb.create_sheet("AI Analysis")
    ws.sheet_view.showGridLines = False

    rows = []
    for a in analyses:
        anomalies     = _parse_json_field(a.get("anomalies", []))
        recommendations = _parse_json_field(a.get("recommendations", []))
        root_causes   = _parse_json_field(a.get("root_causes", []))
        perf          = _parse_json_field(a.get("performance", {}), expected_type=dict)

        top_anom  = anomalies[0]     if anomalies     and isinstance(anomalies[0],     dict) else {}
        top_rc    = root_causes[0]   if root_causes   and isinstance(root_causes[0],   dict) else {}
        top_rec   = recommendations[0] if recommendations and isinstance(recommendations[0], dict) else {}

        rows.append({
            "id": a.get("id"),
            "timestamp": a.get("timestamp", ""),
            "health_status": a.get("health_status", ""),
            "confidence": a.get("confidence"),
            "summary": str(a.get("summary", ""))[:500],
            "anomaly_count": len(anomalies),
            "top_anomaly": top_anom.get("title", ""),
            "top_anomaly_severity": top_anom.get("severity", ""),
            "root_cause_count": len(root_causes),
            "top_root_cause": top_rc.get("issue", ""),
            "recommendation_count": len(recommendations),
            "top_recommendation": top_rec.get("action", ""),
            "bottlenecks": ", ".join(perf.get("bottlenecks", [])) if isinstance(perf, dict) else "",
            "created_at": a.get("created_at", ""),
        })

    df = pd.DataFrame(rows) if rows else pd.DataFrame(
        columns=["id", "timestamp", "health_status", "confidence", "summary",
                 "anomaly_count", "top_anomaly", "top_anomaly_severity",
                 "root_cause_count", "top_root_cause", "recommendation_count",
                 "top_recommendation", "bottlenecks", "created_at"]
    )
    _write_df(ws, df, start_row=1, header_color="1F4E79")

    # Color health status
    status_col = list(df.columns).index("health_status") + 1
    colors = {"HEALTHY": "E2EFDA", "DEGRADED": "FFF3CD", "CRITICAL": "FFE0E0"}
    for row_i in range(2, len(df) + 2):
        status = ws.cell(row=row_i, column=status_col).value
        hex_c = colors.get(status, "FFFFFF")
        for cell in ws[row_i]:
            cell.fill = PatternFill(start_color=hex_c, end_color=hex_c, fill_type="solid")

    _auto_width(ws)
    ws.column_dimensions[get_column_letter(list(df.columns).index("summary") + 1)].width = 60

    # --- Anomalies detail sub-sheet ---
    ws2 = wb.create_sheet("AI Anomalies Detail")
    ws2.sheet_view.showGridLines = False
    anom_rows = []
    for a in analyses:
        ts = a.get("timestamp", "")
        hs = a.get("health_status", "")
        for anom in _parse_json_field(a.get("anomalies", [])):
            if not isinstance(anom, dict):
                continue
            anom_rows.append({
                "analysis_timestamp": ts,
                "health_status": hs,
                "severity": anom.get("severity", ""),
                "title": anom.get("title", ""),
                "detail": str(anom.get("detail", ""))[:300],
                "affected_resources": ", ".join(anom.get("affected_resources", []) or []),
                "evidence": str(anom.get("evidence", ""))[:200],
            })
    df2 = pd.DataFrame(anom_rows) if anom_rows else pd.DataFrame(
        columns=["analysis_timestamp", "health_status", "severity", "title",
                 "detail", "affected_resources", "evidence"]
    )
    _write_df(ws2, df2, start_row=1, header_color="C00000")
    _auto_width(ws2)

    # --- Recommendations detail sub-sheet ---
    ws3 = wb.create_sheet("AI Recommendations")
    ws3.sheet_view.showGridLines = False
    rec_rows = []
    for a in analyses:
        ts = a.get("timestamp", "")
        for rec in _parse_json_field(a.get("recommendations", [])):
            if not isinstance(rec, dict):
                continue
            rec_rows.append({
                "analysis_timestamp": ts,
                "priority": rec.get("priority", ""),
                "action": str(rec.get("action", ""))[:300],
                "reason": str(rec.get("reason", ""))[:200],
                "command": str(rec.get("command", ""))[:200],
            })
    df3 = pd.DataFrame(rec_rows) if rec_rows else pd.DataFrame(
        columns=["analysis_timestamp", "priority", "action", "reason", "command"]
    )
    _write_df(ws3, df3, start_row=1, header_color="375623")
    _auto_width(ws3)
    ws3.column_dimensions["C"].width = 60
    ws3.column_dimensions["E"].width = 60


def _sheet_trends(wb: Workbook, trends: list):
    ws = wb.create_sheet("TSDB Trends")
    ws.sheet_view.showGridLines = False

    rows = []
    for t in trends:
        analysis = _parse_json_field(t.get("analysis", {}), expected_type=dict)
        series_list = analysis.get("series", []) if isinstance(analysis, dict) else []
        for s in series_list:
            labels = s.get("labels", {})
            svc = labels.get("service_name", labels.get("job", "")) if isinstance(labels, dict) else ""
            rows.append({
                "timestamp": t.get("timestamp", ""),
                "query_name": t.get("query_name", ""),
                "description": t.get("description", ""),
                "range_window": t.get("range_window", ""),
                "step": t.get("step", ""),
                "service": svc,
                "data_points": s.get("data_points"),
                "avg": s.get("avg"),
                "min": s.get("min"),
                "max": s.get("max"),
                "latest": s.get("latest"),
                "trend_pct": s.get("trend_pct"),
                "direction": s.get("direction", ""),
                "volatility_cv": s.get("volatility_cv"),
            })

    df = pd.DataFrame(rows) if rows else pd.DataFrame(
        columns=["timestamp", "query_name", "description", "range_window", "step",
                 "service", "data_points", "avg", "min", "max", "latest",
                 "trend_pct", "direction", "volatility_cv"]
    )
    _write_df(ws, df, start_row=1, header_color="1F4E79")

    # Color by direction
    dir_col = list(df.columns).index("direction") + 1
    dir_colors = {"increasing": "FFE0E0", "decreasing": "E2EFDA", "stable": "EBF3FB"}
    for row_i in range(2, len(df) + 2):
        d = ws.cell(row=row_i, column=dir_col).value
        hex_c = dir_colors.get(d, "FFFFFF")
        ws.cell(row=row_i, column=dir_col).fill = PatternFill(
            start_color=hex_c, end_color=hex_c, fill_type="solid"
        )

    _auto_width(ws)

    # Bar chart — latest value per query/service
    if not df.empty and "latest" in df.columns:
        chart = BarChart()
        chart.type = "col"
        chart.title = "Latest Value by Metric Query"
        chart.y_axis.title = "Value"
        pivot = df.groupby("query_name")["latest"].mean().reset_index()
        # Write pivot table for chart
        pivot_start = len(df) + 4
        for c_i, col in enumerate(pivot.columns, 1):
            ws.cell(row=pivot_start, column=c_i, value=col)
        for r_i, row in enumerate(pivot.itertuples(index=False), pivot_start + 1):
            for c_i, val in enumerate(row, 1):
                ws.cell(row=r_i, column=c_i, value=val)
        data_ref = Reference(ws, min_col=2, min_row=pivot_start,
                             max_row=pivot_start + len(pivot))
        cats_ref = Reference(ws, min_col=1, min_row=pivot_start + 1,
                             max_row=pivot_start + len(pivot))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.width = 24
        chart.height = 14
        ws.add_chart(chart, f"A{pivot_start + len(pivot) + 3}")


# ── Main ──────────────────────────────────────────────────────────────────────

def export(base_url: str, since_minutes: int, output_path: str = None):
    print(f"Connecting to backend: {base_url}")
    print(f"Time window: last {since_minutes} minutes\n")

    # Health check
    health = _get(base_url, "/api/health")
    if not health:
        return
    print(f"  Backend status: {health.get('status', 'unknown')}")

    # Fetch all data
    print("  Fetching stats...")
    stats = _get(base_url, "/api/stats") or {}

    print("  Fetching logs...")
    logs_resp = _get(base_url, "/api/logs",
                     params={"since_minutes": since_minutes, "limit": 5000}) or {}
    logs = logs_resp.get("logs", [])
    print(f"    → {len(logs)} log entries")

    print("  Fetching metrics...")
    metrics_resp = _get(base_url, "/api/metrics",
                        params={"since_minutes": since_minutes, "limit": 5000}) or {}
    metrics = metrics_resp.get("metrics", [])
    print(f"    → {len(metrics)} metric entries")

    print("  Fetching traces...")
    traces_resp = _get(base_url, "/api/traces",
                       params={"since_minutes": since_minutes, "limit": 1000}) or {}
    traces = traces_resp.get("traces", [])
    print(f"    → {len(traces)} traces")

    print("  Fetching K8s events...")
    events_resp = _get(base_url, "/api/events",
                       params={"since_minutes": since_minutes, "limit": 1000}) or {}
    events = events_resp.get("events", [])
    print(f"    → {len(events)} events")

    print("  Fetching AI analyses...")
    analysis_resp = _get(base_url, "/api/analysis", params={"limit": 50}) or {}
    analyses = analysis_resp.get("analyses", [])
    print(f"    → {len(analyses)} analyses")

    print("  Fetching TSDB trends...")
    trends_resp = _get(base_url, "/api/tsdb/trends",
                       params={"since_hours": max(1, since_minutes // 60), "limit": 500}) or {}
    trends = trends_resp.get("trends", [])
    print(f"    → {len(trends)} trend records")

    # Build workbook
    print("\nBuilding Excel workbook...")
    wb = Workbook()
    wb.remove(wb.active)

    _sheet_summary(wb, stats, since_minutes)
    _sheet_logs(wb, logs)
    _sheet_metrics(wb, metrics)
    _sheet_traces(wb, traces)
    _sheet_events(wb, events)
    _sheet_analysis(wb, analyses)
    _sheet_trends(wb, trends)

    # Output path
    if output_path is None:
        OUT_DIR.mkdir(parents=True, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = str(OUT_DIR / f"traceflix_data_{ts}.xlsx")

    wb.save(output_path)
    print(f"\nExported → {output_path}")
    print(f"Sheets: {[ws.title for ws in wb.worksheets]}")
    return output_path


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Export TraceFlix live data to Excel")
    parser.add_argument("--url", default=DEFAULT_URL,
                        help=f"Backend API base URL (default: {DEFAULT_URL})")
    parser.add_argument("--since", type=int, default=60,
                        help="Time window in minutes (default: 60)")
    parser.add_argument("--output", default=None,
                        help="Output .xlsx path (default: ml-models/artifacts/traceflix_data_<ts>.xlsx)")
    args = parser.parse_args()
    export(base_url=args.url, since_minutes=args.since, output_path=args.output)
